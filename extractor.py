from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook



PATH = "C:\Program Files (x86)\chromedriver.exe"

driver = webdriver.Chrome()

#year changes and one more place to change on line 73
driver.get("https://www.kerrytaylorauctions.com/archives/?y=2004")

driver.maximize_window()
time.sleep(5)
driver.execute_script("window.scrollBy(0, 500);")
time.sleep(3)

#this will get all "View Results" buttons present on the page
buttons = driver.find_elements(By.XPATH, "//a[contains(text(), 'View Results')]")

#we will click on the buttons one after the other by
# by changing the button number

scroll_distance = 500 # scroll to go slowly down the page

#click on each button
for button in buttons:
    no_of_back = 0
    time.sleep(3)
    button.click() # went inside View details
    no_of_back += 1
    driver.execute_script(f"window.scrollBy(0, 300);")
    time.sleep(3)
    #----------------------------------------------------------------------------------------------------------
    first_lot_link = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, "//p[@class='auction-lot-title']/a"))
        )
    first_lot_link.click() # went inside the first lot
    no_of_back += 1
    time.sleep(3)
    # now we are inside the first lot, we will extract the data
    # and then press next button to go to next lot
    next_arrow_link = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, "//span[@class='pull-right']/a"))
    )
    while next_arrow_link:
        #--------------- extract data here ----------------
        content = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//div[@class='lot lot-details col-sm-6']"))
        )
        # Get the text content of the <div> element
        description = content.text
        workbook = load_workbook("output.xlsx")
        sheet = workbook.active

        # Find the next empty row
        next_row = sheet.max_row + 1
        sheet.cell(row=next_row, column=1,value = description)
        workbook.save("output.xlsx")
        time.sleep(3)
        #  continuously click the next arrow
        next_arrow_link = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//span[@class='pull-right']/a"))
        )
        print(next_arrow_link.get_attribute("href"))
        next_arrow_link.click()
        no_of_back += 1
        time.sleep(5)  
    #----------------------------------------------------------------------------------------------------------
     
    # go back to the previous page
    for i in range(no_of_back):
        driver.back()
        time.sleep(3)
    time.sleep(5)
    scroll_distance += 500
    driver.execute_script(f"window.scrollBy(0, {scroll_distance});")
    time.sleep(5)
    
driver.quit()

